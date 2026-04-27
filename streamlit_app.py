import streamlit as st
import streamlit.components.v1 as components
import anthropic
import base64
import io
import json
import os
import re
from collections import defaultdict

try:
    from streamlit_pdf_viewer import pdf_viewer
    HAS_PDF_VIEWER = True
except ImportError:
    HAS_PDF_VIEWER = False


# ─── API Key ───────────────────────────────────────────────────────────────────
def get_api_key():
    try:
        return st.secrets["ANTHROPIC_API_KEY"]
    except Exception:
        try:
            from dotenv import load_dotenv
            load_dotenv()
        except Exception:
            pass
        return os.environ.get("ANTHROPIC_API_KEY", "")


# ─── Feature Items ─────────────────────────────────────────────────────────────
FEATURE_ITEMS = {
    'シーリングファン':    [11, 12, 13, 15],
    'クッキングコンセント': [34],
    'アンテナ':           [74],
    'ダクトレール':        [7, 8, 9],
    '高機能スイッチ':      [28, 104, 109],
    '調光器照明器具':      [110],
    'ダブル断熱':          [113, 114],
    '自動水栓':            [35, 119],
    '間接照明':            [100, 101, 102, 130],
    '床暖房':              [131],
    'ニッチ':              [2, 86, 87],
    'ガレージ':            [22, 66, 77],
    '乾太くん':            [49],
    '干し姫':              [24, 67],
}

FEATURE_ICONS = {
    'シーリングファン': '🌀', 'クッキングコンセント': '🍳', 'アンテナ': '📡',
    'ダクトレール': '💡', '高機能スイッチ': '🔲', '調光器照明器具': '🎚️',
    'ダブル断熱': '🧱', '自動水栓': '🚿', '間接照明': '🌙',
    '床暖房': '🌡️', 'ニッチ': '🪟', 'ガレージ': '🚗',
    '乾太くん': '👕', '干し姫': '🏠',
}

ITEM_CATEGORY = {}
for n in [1,2,3,4,5,6,7,8,9,10,14,19,20,21,22,33,60,97,98,99,105,106,107,111]:
    ITEM_CATEGORY[str(n)] = '照明・ダウンライト'
for n in [11,12,13,15]:
    ITEM_CATEGORY[str(n)] = 'シーリングファン'
for n in [16,17,18]:
    ITEM_CATEGORY[str(n)] = '人感センサ・ポーチ灯'
for n in [25,26,27,28,29,30,31,32,108,109,110,112]:
    ITEM_CATEGORY[str(n)] = 'スイッチ'
for n in [34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,
          57,58,59,61,62,63,66,67,68,77,79,82,83,84,88,89,
          113,114,117,119,120,121,122,123,124,125,126]:
    ITEM_CATEGORY[str(n)] = 'コンセント・設備'
for n in [50,51,52,53,54,55,56]:
    ITEM_CATEGORY[str(n)] = 'エアコン'
for n in [64,65,69,70,93,94,95,96]:
    ITEM_CATEGORY[str(n)] = '外部設備・給排気'
for n in [71,72,73,74,75,76,80,81,85,86,87,90,91,92,104,115,116]:
    ITEM_CATEGORY[str(n)] = '分電盤・その他設備'
for n in [100,101,102,130]:
    ITEM_CATEGORY[str(n)] = '間接照明'
for n in [23,24,118]:
    ITEM_CATEGORY[str(n)] = '一般的な提案チェック'
for n in [127,128,129]:
    ITEM_CATEGORY[str(n)] = 'インターフォン・リモコン'
for n in [131,132,133,134,135,136]:
    ITEM_CATEGORY[str(n)] = '太陽光・床暖房'

CATEGORY_ORDER = [
    '照明・ダウンライト', 'シーリングファン', '人感センサ・ポーチ灯',
    'スイッチ', 'コンセント・設備', 'エアコン', '外部設備・給排気',
    '分電盤・その他設備', '間接照明', '一般的な提案チェック',
    'インターフォン・リモコン', '太陽光・床暖房',
]


# ─── System Prompt ─────────────────────────────────────────────────────────────
SYSTEM_PROMPT = """あなたは住宅建築の電気図面チェッカーです。
アップロードされた「電気図面（平面図）」と「照明器具配線数量表」のテキスト・注釈情報を読み取り、
以下の全136項目についてチェックしてください。

【チェックの原則】
・テキスト・注釈・シンボルラベルから判断できる項目はOK/NGで判定する
・他の図面（木工事図面・立面図・外観指示図面・プレカット図面等）が必要な項目、
　または視覚的な位置・干渉の確認が必要な項目は「要目視」とする
・図面に該当するキーワード・要素がない項目は「対象外」とする
・独自推測でのエラー報告はしない

【対象外の判定基準（重要）】
「○○がある場合に確認すべき」系の項目は、その設備・器具が図面に存在しない場合は必ず「対象外」とする。
「ない」こと自体はNGではない。以下は特に注意：
・シーリングファン系（11〜13・15）：図面平面にシーリングファンのシンボルが実際に配置されていない場合は「対象外」。凡例への記載のみでは設置ありとみなさない。
・クッキングコンセント（34）：キッチン設備にクッキングコンセントの指示がない場合は「対象外」。クッキングコンセント自体が計画されていない場合はNGではない。
・アンテナ（74）：アンテナ設置の計画・指示がない場合（「アンテナなし」「光テレビ」等、またはそもそも記載なし）は「対象外」。
・乾太くん（49）：乾太くんの記載がない場合は「対象外」。
・先行配管（55）：先行配管の記載がない場合は「対象外」。
・防犯カメラ（85）・プルボックス（76）・スピーカー（90）等：指示がなければ「対象外」。
・太陽光関連（134〜136）：採用しているメーカー・方式のみチェックし、採用していないものは「対象外」。

━━━━━━━━━━━━━━━━━━━━━━━━━━
【照明・ダウンライト】
━━━━━━━━━━━━━━━━━━━━━━━━━━

1. エアコン近くにDL計画がある場合、エアコンが影にならないか注意・確認の注釈があるか

2. ニッチ内にDL設置の場合
   ①木工事図面にΦ数記載の指示があるか
   ②電気図面にΦ数指示があるか（例：※ニッチ内 DL48Φ）

3. ブラケットと干渉物確認の注釈があるか
   ※天付けサッシ・床高変化（小上がり・小下がり）がある場合、照明器具の高さ考慮の指示があるか

4. フットライト等壁埋め込み照明の場合
   ①詳細が図面に記載されているか
   ②ニッチ等薄い壁への設置禁止の注意があるか
   ③コンセント付きフットライトの場合、スイッチ指示が不要であることが確認されているか

5. ペンダントライト・電球下がりタイプの場合
   ①配線位置の指示があるか
   ②FL～下端の高さ指示があるか

6. ブラケットライトの場合、器具の配線位置（高さ・寄り）が記載されているか
   ※階段のブラケットライトは2階平面図に配線位置が記載されることが多い。1階・2階どちらかに記載があれば「OK」とする。

7. ダクトレール採用の場合
   ①設置位置（壁付け・天井付け・梁付け）の記載があるか
   ②何mか記載があるか
   ③器具配線位置の高さ指示（H=1FL+○○等）があるか
   ④吹き抜け壁側面設置の場合、2階図面に記載されているか

8. ダクトレールを化粧梁の上に設置する場合、H=基準（見切り上・梁下端から・H=1FL+○○）が明確に指示されているか

9. アイアン階段がある・梁見せの場合にダクトレールを梁に設置するとき、梁下端～200で指示があるか

10. 階段照明の場合
    ①2階平面図に記載されているか（1階・2階どちらかに記載があればOK）
    ②「壁面用」または「コーナー用」の注釈があるか
       ※「掻き込み用」「引き戸用ブラケット」「階段用」等の表現でも代替可。ブラケットのシンボルがあり高さ指示があればOKとする。
    ③H=1FL+○○形式の高さ指示があるか（H=1FL+xxxx等の記載があればOK）

14. 外部DLは室内と同様のシンボル・指示方法で指示されているか

19. 外壁袖壁に照明・外部コンセントの指示がないか（耐力壁のためNG）

20. 外部DLについて
    ①軒裏が化粧垂木の場合DL設置不可の確認（要目視）
    ②バルコニーDL設置の場合、軒天があるかどうかの確認指示があるか（要目視）

21. バルコニーにDL指示がある場合、軒天の有無を立面図で確認する旨の注釈があるか（要目視）

22. ガレージ内照明について
    ①弊社工事か外構屋かの確認注釈があるか
    ②照明の高さ・設置可能箇所の確認（干渉なし）指示があるか（要目視）

33. 照明設置位置の高さ指示形式が適切か
    ①基準が明確な時：H=○○
    ②基準が明確でない場合（外回り・階段・かさ上げ）：H=○FL+○○

60. 外壁袖壁に照明・外部コンセントの指示がないか（耐力壁のためNG）（19の再確認）

97. 外部照明の検討（バルコニー・ウッドデッキ・勝手口）について、照明デザインの注釈があるか

98. 外部ポスト・表札灯・インターホンを外壁に縦並びで設置する場合、表札灯とインターホンの間が300mm以上あるか

99. 外構照明スイッチを電気図面に指示し「外構○○用スイッチ」と注釈があるか

105. 吹き抜け天井にDL（ダウンライト）の指示がないか（将来の高所対応費がかかるためNG）
     ※スポットライト・ブラケット・ダクトレール・シーリング等のDL以外の照明器具は問題なし（OK）。DL以外が吹き抜けに計上されていてもNGにしない。

106. 2Fに水回りがある場合、直下（1F）の天井点検口とDLの位置関係についてプランナー確認の注釈があるか（要目視）

107. 以下の場合にDL以外（ダクトレール・シーリング・ブラケット等）での提案検討の注釈があるか
     ①吹き抜け天井
     ②バルコニー下・下屋の勾配1寸～3寸・外部に面した梁見せ
     ③階段腰壁が斜めの場合の上部

111. 同じ場所に同種の照明（ブラケット2種・DL2種等）がある場合、どの位置がどの品番かの指示があるか

━━━━━━━━━━━━━━━━━━━━━━━━━━
【シーリングファン】
━━━━━━━━━━━━━━━━━━━━━━━━━━

11. シーリングファン設置の場合
    ①壁からの離隔距離確認の注釈があるか（要目視）
    ②凡例のファンサイズが縮尺で実物サイズに変更されているか（要目視）
    ③延長パイプ検討の記載があるか
    ④スイッチ側に「シーリングファン用スイッチ」と注釈で指示があるか

12. シーリングファンの延長パイプ検討・干渉物（ホスクリーン・物干しバー等）確認の記載があるか（要目視）

13. シーリングファンのスイッチが直接つなげない場合、「シーリングファン用スイッチ」の位置が注釈で指示されているか

15. シーリングファンに昇降機を設置する場合
    ①電気図面に「昇降機付きシーリングファン」と指示があるか
    ②色（ホワイトのみ）の確認記載があるか（要目視）

━━━━━━━━━━━━━━━━━━━━━━━━━━
【人感センサ・ポーチ灯】
━━━━━━━━━━━━━━━━━━━━━━━━━━

16. 人感センサ付きライトが普段触らない位置に計画されているか（要目視）

17. 人感センサ付きDLを連動でご提案している場合、器具に「連動」の文字指示があるか

18. 玄関ポーチ外灯にセンサー付きが提案されているか（要目視）

━━━━━━━━━━━━━━━━━━━━━━━━━━
【スイッチ】
━━━━━━━━━━━━━━━━━━━━━━━━━━

25. トリプルプレート（スイッチ＋コンセント×2）になる場合、図面に注釈指示があるか

26. スイッチ指示の確認
    ①3路・4路の場合、スイッチ赤丸の隣に「3」「4」の数字記載があるか
    ②階をまたぐ3路・4路の場合のみ「○階スイッチとつなぐ」等の注釈が必要。同じ階内で完結する3路・4路はこの注釈不要。
    ③ダブル断熱工法で階段・吹抜付近の3路・4路は外壁面を避けているか（要目視）

27. スイッチ高さを変更する場合
    ①現場一式の変更になっているか（部屋ごとの変更でないか）
    ②「スイッチ：指示のない限り 基本設置位置はH○○○○（芯）」の注意書きがあるか

28. 高機能スイッチ（かってにスイッチ等）を使用する場合
    ①3線式/4線式の場合、図面に品番・親機/子機の指示があるか
    ②センサー感知範囲の確認検討の記載があるか（要目視）

29. スイッチ・コンセント設置壁の仕上がり幅が最低200mm確保されているか（要目視）
    ※小さい壁の場合は他図面への指示が必要

30. UBとびら引き戸の場合、スイッチ設置位置の干渉確認の注釈があるか（要目視）

31. 以下のスイッチ指示が漏れていないか
    ①トイレ換気扇スイッチ
    ②門柱照明スイッチ
    ③外構照明スイッチ
    ④UB照明と換気扇スイッチ（個数確認）

32. スイッチが集中する場合、割り付け指示が平面図に記載されているか

108. 引込戸がある壁にスイッチ・ブラケットの指示がないか（設置不可）

109. 高機能スイッチ（かってにスイッチ・神保熱感スイッチ等）の接続可否確認の記載があるか（要目視）

110. 調光器を提案する場合
     ①照明器具と調光器が同じメーカーか
     ②接続する照明全てが調光可能か

112. 引込戸のところにコンセント指示がないか（設置不可）

━━━━━━━━━━━━━━━━━━━━━━━━━━
【コンセント全般】
━━━━━━━━━━━━━━━━━━━━━━━━━━

34. キッチン設備にクッキングコンセントがない場合、高さ指示（笠木高さとキッチン高さの間）があるか

35. キッチン・洗面台・手洗い自動水栓の場合、自動水栓のシンボル指示があるか

36. コンセント標準設置位置H250以外は、注釈で高さ指示があるか
    ※家具・取付部材との干渉確認も記載があるか（要目視）

37. コンセント横向きに設置する場合、「コンセント横向きに設置」と注釈指示があるか

38. 標準設置位置以外のコンセント高さについて、家具・階段（側板）・ブラケットとの干渉確認の注記があるか（要目視）

39. スイッチ・コンセント設置壁の仕上がり幅が最低200mm確保されているか（要目視）
    ※可動棚レールがある壁面にコンセント提案がないか

40. 畳のかさ上げ天端が300mm未満の場合、コンセント設置不可の確認があるか（要目視）

41. 階段下コンセントの場合、4段目の下から指示されているか（1～3段目はNG）

42. メーカー洗面化粧台設置の場合、コンセント付きか住設図面で確認の注釈があるか（要目視）

43. 造作洗面・造作手洗いのカウンター上にコンセント指示の場合、陶器の高さを考慮した位置指示があるか（要目視）

44. アースコンセントの指示が必要な箇所に指示があるか
    ①トイレ
    ②冷蔵庫
    ③カップボード上
    ④洗面所・脱衣室
    ⑤その他（乾太くん等）
    ※電子レンジ用・洗濯機用は専用回路に「E」がついているか

45. キッチン周りの専用回路について、アースコンセントの指示があるか（要確認）

46. 住宅設備本体に電源が必要な場合、住設図面確認の「※詳細 住宅設備図面参照」の注釈（文字色緑）があるか

47. 食洗機の場合、電気工事書を確認した電圧等の記載（200V等）があるか

48. アップコンセントの場合
    ①寸法線で位置指示があるか（文字指示ではなく）
    ②外壁面からの寸法指示があるか
    ③芯まで指示があるか
    ④色変更がある場合、電気図面左上に品番・カラー指示があるか（文字色3・文字サイズ6pt）

49. 乾太くんがある場合
    ①別紙にガス栓＋アースコンセントの追記があるか
    ②換気計画の注記があるか
    ③電気図面平面上で乾太くんの存在がわかるようになっているか

57. 外部コンセントについて
    ①品番・色の指示があるか
    ②EVコンセントの指示がある場合、急速充電要望の記載があるか
    ③給湯器・プロパンガスから2m離れているか（要目視）
    ④高さ指示がFLからの指示になっているか（GLはNG）
    ⑤立水栓と防水コンセントが455mm以上離れているか（要目視）

58. 給湯器・プロパンガスと防水コンセントの設置位置が2m以上離れているか（要目視）

59. 外部コンセント・EVコンセントの色・品番指示があるか
    ※WK4602WK（ホワイト）・WK4602BK（ブラック）等のパナソニック品番、またはホワイト/ブラックなどの色指定が図面テキストに存在すれば「OK」とする。

61. モデム・ルーター基地の指示
    ①モデムの位置指示があるか
    ②LAN空配管の指示があるか
    ③TVコンセントの指示があるか
    ④2口コンセントの指示があるか

62. 空配管の指示確認
    ①入口と出口に指示があるか
    ②何用か表記があるか（LAN・HDMI等）
    ③複数ある場合、番号が振られているか
    ④別紙指示の場合、平面図にもシンボルと注釈があるか

63. 電動ロールスクリーンを設置する場合
    ①コンセントの指示があるか
    ②「電動ロールスクリーン用コンセント」と注釈があるか
    ③サッシ向かって右側に指示されているか（原則）

66. ガレージにシャッターボックスがある場合
    ①シャッターの設置有無の確認注釈があるか
    ②電動/手動の指示があるか

67. 干し姫（ホシ姫様）がある場合
    ①電動の場合、コンセント指示があるか
    ②壁スイッチタイプの場合、スイッチ位置の指示があるか

68. ブロワの防水コンセント指示があるか（浄化槽がある場合）

77. ガレージ内コンセントについて
    ①弊社工事か外構屋かの確認注釈があるか
    ②設置可能位置と指示があるか
    ③空配管の確認が必要な旨の注釈があるか（要目視）

79. 屋上・お庭がある場合、防水コンセントの提案指示があるか

82. 室外機図面と電気図面のエアコン・室外機位置に変更がある場合、室外機図面の修正指示があるか（要目視）

83. 空配管の指示確認
    ①HDMI空配管の場合「HDMI線用」と注釈があるか（管径が異なるため）
    ②LAN配管のみの場合「LAN空配管」のシンボルが入口・出口に指示があるか
    ③外部空配管の場合、高さ指示があるか

84. LANは実線依頼の場合「LAN入線」、配管のみの場合「LAN空配管」のシンボルが正しく使われているか

88. カウンター・固定棚にコード穴を提案する場合、棚カウンター図面への指示と40Φ等の寸法指示があるか（要目視）

89. 外部空配管の提案の場合、高さ指示があるか

113. ダブル断熱の場合、窓下直下のコンセント提案を避けているか（または注意書きがあるか）（要目視）

114. ダブル断熱の場合、窓下直下のコンセントを極力避けているか（断熱欠損・配管困難）（要目視）

117. 壁掛けテレビの場合、コンセント位置の指示方法が注意事項事例集に準じているか（要目視）

119. キッチン・洗面台・手洗い自動水栓の場合、自動水栓のシンボル指示があるか（35と同）

120. 玄関収納に電動自転車充電用の2口コンセント指示があるか

121. 1階ホールに掃除機動線を考慮した2口コンセント指示があるか

122. 居室のコンセントは対角の位置に指示されているか（建具に寄りすぎていないか）（要目視）

123. リビングテレビスペースに「2口コンセント＋テレビコンセント＋LAN空配管」のセット指示があるか

124. リビングの2口コンセントが最低2箇所提案されているか

125. キッチン背面に「アースコンセント（電子レンジ用）＋2口×2以上」が指示され、高さH=1000（またはカウンター天+100）の指示があるか

126. トイレ（1・2階）・冷蔵庫・洗濯機・電子レンジの5箇所にアースコンセントが指示されているか

━━━━━━━━━━━━━━━━━━━━━━━━━━
【エアコン】
━━━━━━━━━━━━━━━━━━━━━━━━━━

50. エアコンと建具・アウトセット建具ボックスとの干渉確認の記載があるか（要目視）

51. エアコン設置壁面がアウトセット建具ボックス側の場合、幅（約800mm）確保の確認があるか（要目視）

52. エアコンの100V/200Vの指示があるか
    ※リビングに100V・寝室/居室に200Vという通常と異なる提案の場合、目立つ注意書きがあるか

53. スリーブをあける方向の矢印指示があるか（水道・室外機・外観指示図面を参照）（要目視）

54. 弊社受注エアコンが決まっている場合、寸法と品番の資料が別紙として添付されているか

55. 先行配管がある場合、「先行配管あり」「ダクトカバー：（色）」の注釈指示があるか

56. エアコン近くにPS（点検口）等干渉物がないか確認の記載があるか（要目視）

━━━━━━━━━━━━━━━━━━━━━━━━━━
【外部設備・給排気・ダクト】
━━━━━━━━━━━━━━━━━━━━━━━━━━

64. 電動シャッター・電動ユニットの指示漏れがないか（サッシ図面確認、窓横に指示）（要目視）
    ※電動シャッターがある場合、同壁面にコンセント/スイッチの提案があるか

65. 電動シャッターがある場合、同じ壁面にコンセントかスイッチの提案があるか（要目視）

69. キッチンダクトの指示
    ①ダクト抜き方向に干渉物がないか（要目視）
    ②キッチンダクト抜き方向の指示が電気図面にあるか
    ③同時給排気型の場合、2本分の指示があるか
    ④外部ダクトフード色変更がある場合、電気図面左上に変更の記載があるか（文字色黄緑・文字大きさ5pt）

70. 城陽市の場合、キッチンフードダクトにスパイラルダクトの指示があるか

93. 給湯器・換気扇・給気口の指示について
    ①設計士から引き継いだ内容で位置指示があるか
    ②設置位置がメンテナンス可能なスペースを確保しているか（要目視）
    ③給気口がエアコン等センサー家電から離れた位置にあるか（要目視）

94. 天井付け換気扇の場合、位置指示とダクト構造回避の指示があるか
    ※壁付け換気扇の場合、開口指示があるか
    ※スクエア窓等換気ブレスが付かないサッシの場合、給気口の設置位置・高さ指示があるか

95. 換気扇の変更予備費（54,000円/6箇所）が計上されているか（数量表チェック）
    ※この物件で換気扇変更予備費が不要・計上なしの場合は「対象外」とする。

96. 干渉物（ガーデンパン・竪樋・室外機・ブラケット・外部コンセント等）の確認の記載があるか（要目視）
    ※窓モール等の装飾がある場合、給気口スリーブとの干渉確認があるか（要目視）

━━━━━━━━━━━━━━━━━━━━━━━━━━
【防犯カメラ】
━━━━━━━━━━━━━━━━━━━━━━━━━━

85. 防犯カメラの場合
    ①木工事図面への下地指示の記載があるか（要目視）
    ②以下のいずれかの注釈があるか
       ・無線：「防犯カメラ用電源」
       ・有線：「防犯カメラ用」＋空配管指示
       ・PoE：「防犯カメラ用」＋PoE給電指示
    ③軒天設置の場合、壁面からの寄り（100mm以上）の記載があるか

━━━━━━━━━━━━━━━━━━━━━━━━━━
【ニッチ】
━━━━━━━━━━━━━━━━━━━━━━━━━━

86. ニッチ内コンセントの場合
    ①奥行が確保できているか確認の注釈があるか（要目視）
    ②ニッチ寸法変更がある場合、設計士確認の記載があるか（要目視）

87. ニッチ内コンセントの場合（86と同様の再確認）
    ①奥行確保の確認があるか（要目視）
    ②ニッチ寸法変更時、設計士確認の記載があるか（要目視）

━━━━━━━━━━━━━━━━━━━━━━━━━━
【分電盤・その他設備】
━━━━━━━━━━━━━━━━━━━━━━━━━━

71. 分電盤の設置位置が以下の禁止箇所になっていないか（設置スペース350mm必要）
    ①外壁面
    ②耐力壁
    ③階段/梁見せに面している壁
    ④UB建具上
    ⑤H20を超える建具上
    ⑥アウトセットレールがある側の建具上
    ⑦干渉物（ホスクリーン・ガラスブロック等）との確認（要目視）

72. 各種リモコンの電気図面指示があるか
    ①インターフォンリモコン（外壁設置の場合H1450・寄り記載）
    ②給湯器リモコン
    ③太陽光リモコン

73. インターフォンが外壁に付く場合、設置位置（土間仕上げH1450・寄り）の指示があるか
    ※外壁にインターフォンが設置されない場合は「対象外」とする。

74. アンテナ設置の場合、図面右下確認事項欄に記載（マスプロアンテナ・BSアンテナ有無・カラー）があるか
    ※光テレビの場合、モデム基地に「4口コンセント+TV端子」の指示があるか

75. NTTフレッツ光（ソフトバンク光）でTV視聴の場合、モデム基地に「4口コンセント+TV端子」の指示があるか

76. プルボックス指定位置がある場合、シンボル指示と確認事項欄「プルボックス位置指定：有」の記載があるか
    ※現場合わせの場合は確認事項欄に斜線があるか

80. 文字・数字が逆さ文字になっていないか（要目視）
    ※シンボルが集中して読み取りづらい場合、平面図外への指示があるか（要目視）

81. 以下の場合に別紙が作成されているか
    ①造作別紙がある場合
    ②タイル施工がある場合
    ③造作カップボードがある場合
    ④外壁面にTVコンセント＋他の指示がある場合
    ⑤色が関わる設置物がある場合
    ⑥クロス見切り材とスイッチが干渉する可能性がある場合
    ⑦インターホン・太陽光リモコン・給湯器の位置（必ず別紙）

90. スピーカー設置の場合
    ①設置位置（直置き・壁付け・天井付け）の確認があるか
    ②壁/天井付けの場合、木工事図面への下地指示があるか（要目視）
    ③天井付けの場合、梁下への設置指示がないか（梁下はNG）

91. 腰壁・アイアン等で壁仕上がりが天井まで到達しない場合、図面に指示があるか（文字サイズ2.5・文字色黒）

92. 別紙への参照指示「※詳細 別紙○ 参照」の注釈（文字色黄緑・点線2・文字大きさ5pt）があるか

116. 特殊な電気製品（電気釜・ガレージリフト等）がある場合、必要電源の種類・ブレーカー有無・現地調査要否の確認記載があるか（要目視）

━━━━━━━━━━━━━━━━━━━━━━━━━━
【間接照明】
━━━━━━━━━━━━━━━━━━━━━━━━━━

100. 間接照明がある場合
     ①電気図面に設置方法（壁付け/天井付け・梁上・玄関収納下等）の注釈指示があるか
     ②積算明細表で正しく計上されているか確認の記載があるか（要目視）

101. 間接照明の検討・注意点について
     ①有効寸法200mm以上確保の確認があるか（要目視）
     ②垂れ壁は最低85mm以上あるか（要目視）
     ③玄関框（180mm）への間接照明指示がないか（施工困難なためNG）

102. 間接照明がある箇所に板張りとクロス張りが混在する場合、断面図上にクロスと板張りの範囲記載があるか（要目視）

130. 間接照明の設置方法注釈（配線位置・壁/天井付け）があるか
     ※レンジフードとの離隔距離確認（天井高を下げる場合）の記載があるか（要目視）
     ※玄関框180mmへの指示がないか

━━━━━━━━━━━━━━━━━━━━━━━━━━
【一般的な提案チェック】
━━━━━━━━━━━━━━━━━━━━━━━━━━

23. 設計士から引き継いだ図面未記載設備（カップボード・キッチンフード・吊り戸等）の寸法・設置位置が図面に記載されているか（要目視）

24. ホシ姫様タイプの物干しがある場合、天井照明（DL・引掛けシーリング）との干渉確認の記載があるか（要目視）

118. モデルの場合、テレビコンセントが全室・引掛けシーリングが全室提案されているか

━━━━━━━━━━━━━━━━━━━━━━━━━━
【太陽光・床暖房・床下エアコン】
━━━━━━━━━━━━━━━━━━━━━━━━━━

131. 床暖房がある場合
     ①設置範囲・リモコン位置が電気図面に指示されているか
     ②床図面に「床暖房施工あり、寸法指示図参照」の指示があるか（要目視）
     ③床暖房施工範囲に床下点検口・アップコンセントの干渉がないか（要目視）

132. 床下エアコンの別紙にコンセント位置が明確に記載されているか

133. 電動ロールスクリーンのコンセントが原則サッシ向かって右側に指示されているか
     ※左側設置の場合、スライド確認の記載があるか

134. 建て得（リクシル×テプコ）の場合
     ①パワコン1台に対して非常用コンセント1口指示があるか
     ②分電盤隣に送信用ユニット・一括制御リモコン（「パワコン用リモコン」）の指示があるか

135. ナカザワ建販（エクソル）の場合
     ①パワコン1台に対して非常用コンセント2口指示があるか
     ②太陽光ブレーカー（40A×1込）分電盤の指示があるか

136. 太陽光共通の指示
     ①電気図面にパワコンの位置と台数の指示があるか
     ②パワコンが北面・西面の日陰で風通しの良い位置に指示されているか（要目視）
     ③パワコン設置禁止箇所（梁見せ壁・吹抜け壁・階段壁・給湯器上）への指示がないか

━━━━━━━━━━━━━━━━━━━━━━━━━━
【インターフォン・アンテナ・リモコン類】
━━━━━━━━━━━━━━━━━━━━━━━━━━

127. インターフォンリモコンがリビング扉の近く（リビング側）に設置されているか（要目視）

128. 給湯器リモコンがキッチン壁の近くに設置されているか（要目視）

129. モデム・LAN空配管・TVコンセント・4口コンセントの指示が1階中心の収納内に指示されているか

━━━━━━━━━━━━━━━━━━━━━━━━━━
【その他】
━━━━━━━━━━━━━━━━━━━━━━━━━━

104. かってにスイッチや標準外スイッチ等の特殊商品を使用する場合、図面に注意事項の記載があるか

115. 店舗併用住宅で天井埋込エアコンを利用の場合、点検口設置位置の確認注釈があるか（要目視）

━━━━━━━━━━━━━━━━━━━━━━━━━━
【出力フォーマット】
━━━━━━━━━━━━━━━━━━━━━━━━━━
必ず以下のJSON形式のみで出力してください。JSONの前後に説明文・マークダウン・コードフェンスは不要です。

{
  "property": "（読み取れた物件名。不明な場合は「不明」）",
  "items": [
    {
      "id": "1",
      "title": "エアコン近くのDL干渉確認",
      "status": "OK",
      "detail": "エアコンとDLの注釈確認",
      "evidence": "「※エアコン付近DL影注意」 図面右下注記付近"
    }
  ]
}

statusの値は必ず以下のいずれかにしてください：
- "OK"   : チェックして問題なし
- "NG"   : エラー・記載漏れ・ルール違反あり（detailに具体的な問題内容を記載）
- "要目視": AIでは判定不能（視覚確認・他図面確認が必要）
- "対象外": 図面に該当するキーワード・要素がなく、チェック不要

上記の全項目を必ず含めること（id: "1"〜"136"）。
detailは1〜2文で簡潔に記載すること。
evidenceは判定の根拠となった図面内の注釈・シンボル・テキストを引用し、可能なら位置情報（例:「右下凡例」「LDK付近」「2階平面図」）も含めて記載すること。
- "OK"の場合: 該当する注釈・テキスト・シンボルを引用（例:「『※エアコン付近DL影注意』 右下注記」）
- "NG"の場合: 不足している記述や問題箇所を具体的に（例:「『シーリングファン用スイッチ』の注釈が見当たらない」）
- "要目視"の場合: 視覚確認が必要な対象・場所（例:「平面図中央の階段付近、ブラケット高さ要確認」）
- "対象外"の場合: 簡潔に（例:「該当設備（シーリングファン）なし」）"""


# ─── DXF Parser ────────────────────────────────────────────────────────────────
def parse_dxf(file_bytes):
    content = file_bytes.decode('utf-8', errors='ignore')
    lines = [l.rstrip() for l in content.split('\n')]

    def decode_text(s):
        return s.replace('\\U', '')

    texts, inserts = [], []
    i = 0
    while i < len(lines) - 1:
        code = lines[i].strip()
        val = lines[i+1].strip()
        if code == '0' and val in ('TEXT', 'MTEXT', 'INSERT'):
            etype = val
            layer, text, block = '', '', ''
            i += 2
            while i < len(lines) - 1:
                c = lines[i].strip()
                v = lines[i+1].strip()
                if c == '0':
                    break
                if c == '8': layer = decode_text(v)
                if c == '1': text = decode_text(v)
                if c == '2' and etype == 'INSERT': block = decode_text(v)
                i += 2
            if etype in ('TEXT', 'MTEXT') and text.strip():
                texts.append((layer, text))
            if etype == 'INSERT' and block and not block.startswith('*'):
                inserts.append((layer, block))
            continue
        i += 2
    return texts, inserts


def format_dxf_for_claude(texts, inserts):
    layer_texts = defaultdict(list)
    for layer, text in texts:
        layer_texts[layer].append(text)

    out = ['=== 電気図面 テキスト・シンボル情報（DXF解析）===', '']
    priority = ['??? 4', '??? 20', '??? 3']
    for layer in priority:
        if layer not in layer_texts:
            continue
        label = {'??? 4': '図面内注釈', '??? 20': 'アース指示', '??? 3': '凡例・シンボル注釈'}.get(layer, layer)
        out.append(f'【{label}】')
        for t in layer_texts[layer]:
            if t.strip():
                out.append(f'  {t}')
        out.append('')

    for layer, txts in layer_texts.items():
        if layer in priority or layer == '??? 1':
            continue
        out.append(f'【レイヤー {layer}】')
        for t in txts:
            if t.strip():
                out.append(f'  {t}')
        out.append('')

    seen = set()
    symbols = []
    for _, block in inserts:
        clean = re.sub(r'_SfigorgFlag_\d+', '', block).strip()
        clean = re.sub(r'\?+', '', clean).strip()
        if clean and clean not in seen:
            symbols.append(clean)
            seen.add(clean)
    if symbols:
        out.append('【使用シンボル一覧】')
        for s in symbols:
            out.append(f'  {s}')
    return '\n'.join(out)


# ─── XLSX Parser ───────────────────────────────────────────────────────────────
def parse_xlsx(file_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    out = ['=== Excelファイル（照明器具配線数量表）===', '']
    for sheet_name in wb.sheetnames:
        if not (('照明' in sheet_name and '数量' in sheet_name) or '照明チェック' in sheet_name):
            continue
        ws = wb[sheet_name]
        out.append(f'--- シート: {sheet_name} ---')
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip() and str(c).strip() != 'None']
            if cells:
                out.append(' | '.join(cells))
                row_count += 1
                if row_count >= 120:
                    break
        out.append('')
    return '\n'.join(out)


# ─── JSON Parser ───────────────────────────────────────────────────────────────
def parse_json_response(text):
    text = text.strip()
    cleaned = text
    if text.startswith('```'):
        first_newline = text.find('\n')
        if first_newline != -1:
            cleaned = text[first_newline + 1:]
        last_fence = cleaned.rfind('```')
        if last_fence != -1:
            cleaned = cleaned[:last_fence]
        cleaned = cleaned.strip()
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass
    for target in [cleaned, text]:
        start = target.find('{')
        end = target.rfind('}')
        if start != -1 and end != -1 and end > start:
            try:
                return json.loads(target[start:end + 1])
            except json.JSONDecodeError:
                pass
    # Repair attempt: response truncated mid-JSON. Close at last complete item.
    for target in [cleaned, text]:
        start = target.find('{')
        if start == -1:
            continue
        candidate = target[start:]
        items_pos = candidate.find('"items"')
        if items_pos == -1:
            continue
        last_complete = candidate.rfind('},', items_pos)
        if last_complete == -1:
            last_complete = candidate.rfind('}', items_pos)
            if last_complete == -1:
                continue
        for closing in ('\n  ]\n}', ']}', '}\n  ]\n}'):
            try:
                return json.loads(candidate[:last_complete + 1] + closing)
            except json.JSONDecodeError:
                continue
    return None


# ─── Check Runner ──────────────────────────────────────────────────────────────
def run_check(drawing_bytes, drawing_name, table_bytes, table_name, selected_features):
    api_key = get_api_key()
    if not api_key:
        st.error("APIキーが設定されていません。Streamlit CloudのSecretsにANTHROPIC_API_KEYを設定してください。")
        return None

    excluded_items, excluded_names = [], []
    for feature, items in FEATURE_ITEMS.items():
        if feature not in selected_features:
            excluded_items.extend(items)
            excluded_names.append(feature)

    exclusion_note = ''
    if excluded_items:
        ids_str = '・'.join(str(i) for i in sorted(set(excluded_items)))
        exclusion_note = (
            f'\n\n【この物件の対象外設備】\n'
            f'以下の設備はこの物件に存在しないため、関連項目（{ids_str}）は必ず「対象外」として回答してください。\n'
            f'対象外設備: {"、".join(excluded_names)}'
        )

    drawing_lower = drawing_name.lower()
    table_lower = table_name.lower()
    drawing_is_dxf = drawing_lower.endswith('.dxf')
    table_is_xlsx = table_lower.endswith('.xlsx') or table_lower.endswith('.xls')

    if table_is_xlsx:
        table_text = parse_xlsx(table_bytes)
        table_content = {"type": "text", "text": f"以下は照明器具配線数量表（Excelファイル）から抽出したデータです。\n\n{table_text}"}
    else:
        table_data = base64.standard_b64encode(table_bytes).decode('utf-8')
        table_content = {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": table_data}, "title": "照明器具配線数量表"}

    if drawing_is_dxf:
        texts, inserts = parse_dxf(drawing_bytes)
        drawing_text = format_dxf_for_claude(texts, inserts)
        content = [
            {"type": "text", "text": f"以下は電気図面のDXFファイルから抽出したテキスト・シンボル情報です。\n\n{drawing_text}"},
            table_content,
            {"type": "text", "text": f"上記の電気図面情報と照明器具配線数量表をチェックしてください。指定のJSON形式で全項目の結果を出力してください。{exclusion_note}"},
        ]
    else:
        drawing_data = base64.standard_b64encode(drawing_bytes).decode('utf-8')
        content = [
            {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": drawing_data}, "title": "電気図面（平面図）"},
            table_content,
            {"type": "text", "text": f"上記の電気図面と照明器具配線数量表をチェックしてください。指定のJSON形式で全項目の結果を出力してください。{exclusion_note}"},
        ]

    try:
        client = anthropic.Anthropic(api_key=api_key)
        message = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=32000,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": content}],
        )
        raw = message.content[0].text
        parsed = parse_json_response(raw)
        if parsed is None:
            st.session_state['debug_raw_response'] = raw
            st.error(f"解析失敗（全長:{len(raw)}文字）。Claude APIのレスポンスが不正です。")
            return None
        st.session_state.pop('debug_raw_response', None)
        return parsed
    except anthropic.AuthenticationError:
        st.error("APIキーが無効です。正しいAPIキーを設定してください。")
        return None
    except anthropic.RateLimitError:
        st.error("APIの利用制限に達しました。しばらく待ってから再試行してください。")
        return None
    except Exception as e:
        st.error(f"エラーが発生しました: {str(e)}")
        return None


# ─── PDF Preview ───────────────────────────────────────────────────────────────
def render_pdf_preview(pdf_bytes, height=780):
    if HAS_PDF_VIEWER:
        pdf_viewer(pdf_bytes, height=height, width="100%")
    else:
        base64_pdf = base64.b64encode(pdf_bytes).decode('utf-8')
        pdf_html = f'<iframe src="data:application/pdf;base64,{base64_pdf}" width="100%" height="{height}" style="border:1px solid #e2e8f0; border-radius:8px;"></iframe>'
        components.html(pdf_html, height=height + 10, scrolling=False)


# ─── Result Rendering ──────────────────────────────────────────────────────────
STATUS_CFG = {
    'OK':    {'color': '#16a34a', 'bg': '#f0fdf4', 'border': '#86efac', 'label': 'OK'},
    'NG':    {'color': '#dc2626', 'bg': '#fef2f2', 'border': '#fca5a5', 'label': 'NG'},
    '要目視': {'color': '#d97706', 'bg': '#fffbeb', 'border': '#fcd34d', 'label': '要目視'},
    '対象外': {'color': '#94a3b8', 'bg': '#f8fafc', 'border': '#e2e8f0', 'label': '対象外'},
}

def render_card(item):
    cfg = STATUS_CFG.get(item.get('status', '対象外'), STATUS_CFG['対象外'])
    id_ = str(item.get('id', ''))
    title = str(item.get('title', ''))
    detail = str(item.get('detail', ''))
    evidence = str(item.get('evidence', '')).strip()
    evidence_html = ''
    if evidence:
        evidence_html = (
            f'<div style="font-size:11px;color:#64748b;margin-top:6px;'
            f'padding:5px 8px;background:#f8fafc;border-radius:4px;'
            f'border-left:2px solid #cbd5e1;line-height:1.5;">'
            f'📌 <span style="color:#475569;">{evidence}</span></div>'
        )
    html = f"""<div style="border-left:4px solid {cfg['color']};background:{cfg['bg']};
        border:1px solid {cfg['border']};border-left:4px solid {cfg['color']};
        border-radius:8px;padding:10px 14px;margin-bottom:6px;
        display:flex;align-items:flex-start;gap:10px;">
        <span style="font-family:monospace;font-size:11px;color:#94a3b8;
            min-width:2rem;padding-top:2px;">{id_}</span>
        <div style="flex:1;">
            <div style="font-weight:600;font-size:13px;color:#374151;">{title}</div>
            <div style="font-size:12px;color:#6b7280;margin-top:3px;">{detail}</div>
            {evidence_html}
        </div>
        <span style="background:{cfg['color']};color:white;font-size:11px;font-weight:bold;
            padding:2px 8px;border-radius:9999px;white-space:nowrap;">{cfg['label']}</span>
    </div>"""
    st.markdown(html, unsafe_allow_html=True)


def render_results(data):
    property_name = data.get('property', '不明')
    items = data.get('items', [])

    counts = {'OK': 0, 'NG': 0, '要目視': 0, '対象外': 0}
    for item in items:
        if item.get('status') in counts:
            counts[item['status']] += 1

    st.subheader(f"📋 {property_name}")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("🔴 NG", counts['NG'])
    c2.metric("🟡 要目視", counts['要目視'])
    c3.metric("🟢 OK", counts['OK'])
    c4.metric("⚪ 対象外", counts['対象外'])

    # NG items first
    ng_items = [i for i in items if i.get('status') == 'NG']
    if ng_items:
        st.markdown(f"### 🔴 NG項目（{len(ng_items)}件）— 要対応")
        for item in ng_items:
            render_card(item)
        st.markdown("---")

    # All items by category
    st.markdown("### 📋 全項目チェック結果")
    grouped = defaultdict(list)
    for item in items:
        cat = ITEM_CATEGORY.get(str(item.get('id', '')), 'その他')
        grouped[cat].append(item)

    for cat in CATEGORY_ORDER + ['その他']:
        cat_items = grouped.get(cat)
        if not cat_items:
            continue
        ng_c = sum(1 for i in cat_items if i.get('status') == 'NG')
        label = f"**{cat}**（{len(cat_items)}件" + (f"　🔴NG:{ng_c}" if ng_c else "") + "）"
        with st.expander(label, expanded=(ng_c > 0)):
            for item in sorted(cat_items, key=lambda x: int(x.get('id', 0)) if str(x.get('id', '')).isdigit() else 999):
                render_card(item)

    # Copy text button
    st.markdown("---")
    copy_text = f"{property_name}\n\n"
    for item in items:
        copy_text += f"[{item.get('status','')}] {item.get('id','')} {item.get('title','')}\n  {item.get('detail','')}\n"
        ev = str(item.get('evidence', '')).strip()
        if ev:
            copy_text += f"  📌 {ev}\n"
    st.download_button("📋 結果をテキストでダウンロード", copy_text, file_name="check_result.txt", mime="text/plain")


# ─── Main ──────────────────────────────────────────────────────────────────────
def main():
    st.set_page_config(page_title="電気図面チェッカー", page_icon="⚡", layout="wide")

    has_pdf_preview = bool(st.session_state.get('result') and st.session_state.get('drawing_pdf_bytes'))
    max_w = 1400 if has_pdf_preview else 860
    st.markdown(f"""
    <style>
    .block-container {{ max-width: {max_w}px; padding-top: 2rem; }}
    div[data-testid="stHorizontalBlock"] {{ align-items: flex-start; }}
    </style>
    """, unsafe_allow_html=True)

    st.title("⚡ 電気図面チェッカー")
    st.caption("電気図面と照明器具配線数量表をアップロードして自動チェックします")

    # ── File Uploads ──────────────────────────────────────────────────────
    col1, col2 = st.columns(2)
    with col1:
        drawing_file = st.file_uploader("📐 電気図面（平面図）", type=["pdf", "dxf"], help="PDF または DXF形式")
        if drawing_file:
            ext = drawing_file.name.rsplit('.', 1)[-1].upper()
            if ext == 'DXF':
                st.success(f"✅ {drawing_file.name}　**DXF（高精度）**")
            else:
                st.info(f"✅ {drawing_file.name}　PDF")
    with col2:
        table_file = st.file_uploader("📊 照明器具配線数量表", type=["pdf", "xlsx", "xls"], help="PDF または Excel形式")
        if table_file:
            ext = table_file.name.rsplit('.', 1)[-1].upper()
            if ext in ('XLSX', 'XLS'):
                st.success(f"✅ {table_file.name}　**Excel（高精度）**")
            else:
                st.info(f"✅ {table_file.name}　PDF")

    # ── Feature Selection ─────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("**🏠 この物件に含まれる設備・オプション**")
    st.caption("チェックした設備のみチェック対象になります（未チェック＝対象外）")

    selected_features = []
    cols = st.columns(3)
    for i, feature in enumerate(FEATURE_ITEMS.keys()):
        with cols[i % 3]:
            icon = FEATURE_ICONS.get(feature, '')
            if st.checkbox(f"{icon} {feature}", key=f"feat_{feature}"):
                selected_features.append(feature)

    # ── Submit ────────────────────────────────────────────────────────────
    st.markdown("---")
    if st.button("🔍 チェックを開始する", type="primary", use_container_width=True):
        if not drawing_file:
            st.error("電気図面（PDFまたはDXF）をアップロードしてください。")
        elif not table_file:
            st.error("照明器具配線数量表（PDFまたはExcel）をアップロードしてください。")
        else:
            drawing_bytes = drawing_file.read()
            table_bytes = table_file.read()
            with st.spinner("AIが図面を解析中です... (30秒〜1分ほどかかります)"):
                result = run_check(
                    drawing_bytes, drawing_file.name,
                    table_bytes, table_file.name,
                    selected_features,
                )
            if result:
                st.session_state['result'] = result
                if drawing_file.name.lower().endswith('.pdf'):
                    st.session_state['drawing_pdf_bytes'] = drawing_bytes
                else:
                    st.session_state.pop('drawing_pdf_bytes', None)
                st.rerun()

    # Show raw response on parse failure for debugging
    if st.session_state.get('debug_raw_response'):
        with st.expander("🔍 デバッグ: Claude API レスポンス（解析失敗時の生データ）"):
            raw = st.session_state['debug_raw_response']
            st.caption(f"全長: {len(raw)} 文字")
            st.text(raw[:500] + ('\n\n...（中略）...\n\n' + raw[-500:] if len(raw) > 1000 else ''))

    # ── Results ───────────────────────────────────────────────────────────
    if st.session_state.get('result'):
        st.markdown("---")
        pdf_bytes = st.session_state.get('drawing_pdf_bytes')
        if pdf_bytes:
            col_pdf, col_results = st.columns([1, 1], gap="medium")
            with col_pdf:
                st.markdown("**📐 図面プレビュー**")
                render_pdf_preview(pdf_bytes)
            with col_results:
                render_results(st.session_state['result'])
        else:
            render_results(st.session_state['result'])
        if st.button("🔄 リセット"):
            st.session_state.pop('result', None)
            st.session_state.pop('drawing_pdf_bytes', None)
            st.rerun()

    st.markdown("---")
    st.caption("Powered by Claude API（claude-sonnet-4-6）")


if __name__ == "__main__":
    main()
